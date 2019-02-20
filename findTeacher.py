import sys
from random import randint
from time import sleep
from openpyxl import load_workbook
import urllib
import requests
from bs4 import BeautifulSoup

#請求指定一個請求來模擬chrome瀏覽器
global headers
headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.75 Safari/537.36'}

if sys.platform.startswith('win'):
  wb = load_workbook(filename = 'teacherList.xlsx')
elif sys.platform.startswith('dar'):
  wb = load_workbook(filename='/Users/williamchuang/PycharmProjects/FindTeacher/teacherList.xlsx')
else:
  print("OS system not support (only windows and mac)")
print(sys.platform)
print(wb.sheetnames)


ws = wb.active

i = 2
while i < 10:
   sName = ws.cell(row=i, column=6).value
   dName = ws.cell(row=i, column=7).value
   tName = ws.cell(row=i, column=8).value
   i += 1

   #搜尋網址
   #url = 'https://www.google.com/search?q=%E4%B8%AD%E8%8F%AF%E5%A4%A7%E5%AD%B8%E5%BC%B5%E8%80%80%E6%96%87'
   #sName = urllib.parse.quote("中華大學")
   #dName = urllib.parse.quote("科技管理學系")
   #tName = urllib.parse.quote("張耀文")
   # print(sName, dName, tName)

   url = 'https://www.google.com/search?q='+sName+dName+tName
   # html_tree = lxml.parse(url)
   # print(html_tree)

   html = requests.get(url, headers=headers)
   soup = BeautifulSoup(html.text, 'html.parser')
   st1 = soup.find_all('div',class_='r')

   print("start search with 3~5 sleep---------------")
   sleep(randint(3, 5))

   for st2 in st1:
     st3 = st2.find('a')
     print('name: ' + st3.text)
     # x = st3.get('href')
     # y = x[x.index('http'): x.index('&')]
     teacherURL = st3['href']
     print('link: ' + teacherURL)

     teacherHtml = requests.get(teacherURL, headers=headers)
     teacherHtml.encoding = 'utf-8'
     teacherSoup = BeautifulSoup(teacherHtml.text, 'html.parser')

     #去除javascript
     for s in teacherSoup('script'):
        s.extract()

     #print(teacherSoup.prettify('utf-8'))
     keyMan = ["授課", "學歷", "專長", "mail", "電話", "郵箱"]
     foundIt = False
     foundCount = 0

        #if link.find("學 歷")>-1 or link.find("授課領域")>-1 or link.find("專長")>-1 or link.find("職        稱")>-1 or link.find("姓名")>-1 or link.find("電話")>-1:
        #if "授課" in link or "學歷" in link or "專長" in link or "mail" in link or "電話" in link or "郵箱" in link:
        # if any( c in link for c in ("授課", "學歷", "專長", "mail", "電話", "郵箱") ) :
     for link in teacherSoup.find_all(text = True):
         if any( c in link for c in keyMan ) :
           foundIt = True
           foundCount += 1
           print("[", foundCount, "]", link, "-end \n")
         else:
           continue

     if foundIt:
        break
     else:
        continue

#print(arr)





   #from urllib.request import urlopen
   #from bs4 import BeautifulSoup
   #import re
   #html = urlopen("http://en.wikipedia.org/wiki/Kevin_Bacon")
   #bsObj = BeautifulSoup(html, features="lxml")
   #for link in bsObj.find("div", {"id":"bodyContent"}).findAll("a",href=re.compile("^(/wiki/)((?!:).)*$")):
   #    if 'href' in link.attrs:
   #        print(link.attrs['href'])


   # google_url = 'https://www.google.com.tw/search'
   # query = input('輸入關鍵字')
   # my_params = {'q': query}
   # url = 'https://www.google.com/'
   # headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.75 Safari/537.36'}
   # req = urllib.request.Request(url, headers=headers)
   # x = urllib.request.urlopen(req).read()
   # print(x)
   # def getHtml(url):
   #     page = urllib.request.urlopen(url)
   #     html = page.read()
   #     return html

   # htmls = getHtml('https://www.google.com/search?q=%E4%B8%AD%E8%8F%AF%E5%A4%A7%E5%AD%B8%E5%BC%B5%E8%80%80%E6%96%87')
   # print(htmls)


   #r = requests.get(google_url, params = my_params)


   # if r.status_code == requests.codes.ok:
   #     soup = str(BeautifulSoup(r.text, 'html.parser'))
   #     selector = etree.HTML(soup)
   #     links = selector.xpath('//div[@class="r"]/a/h3')

   #     #items = soup.select('div.g > h3.r > a[href^="/url"]')
   #     print(links)

